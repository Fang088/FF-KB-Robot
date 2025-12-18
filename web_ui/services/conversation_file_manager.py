"""
对话文件管理器 - 处理对话中上传的文件

功能：
1. 保存上传的文件到临时目录
2. 提取文件内容（支持多种格式）
3. 管理文件生命周期（TTL）
4. 清理过期的临时文件
5. 文件去重（基于哈希）

作者: FF-KB-Robot Team
"""

import logging
import json
from pathlib import Path
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from dataclasses import dataclass, asdict
import uuid
import io

logger = logging.getLogger(__name__)


class ConversationFileManagerError(Exception):
    """对话文件管理异常"""
    pass


@dataclass
class UploadedFileInfo:
    """上传文件的信息"""
    file_id: str              # 文件的唯一ID（哈希）
    filename: str             # 原始文件名
    file_type: str            # 文件类型（txt, pdf, image 等）
    file_size: int            # 文件大小（字节）
    upload_time: str          # 上传时间（ISO 格式）
    content_preview: str      # 内容预览或缩略图（base64 或文本）
    metadata: Dict[str, Any]  # 额外的元数据

    def to_dict(self) -> Dict[str, Any]:
        """转换为字典"""
        return asdict(self)


class ConversationFileManager:
    """
    对话文件管理器 - 管理对话中上传的临时文件

    职责：
    1. 保存和管理上传的文件
    2. 提取文件内容
    3. TTL 管理
    4. 清理过期文件
    5. 文件去重
    """

    def __init__(
        self,
        temp_base_dir: str,
        max_file_size_mb: int = 100,
        max_file_content_length: int = 5000,
        max_pdf_pages: int = 3
    ):
        """
        初始化对话文件管理器

        Args:
            temp_base_dir: 临时文件基础目录
            max_file_size_mb: 单个文件最大大小（MB）
            max_file_content_length: 提取的文件内容最大长度（字符）
            max_pdf_pages: PDF 文件最多提取的页数
        """
        self.temp_base_dir = Path(temp_base_dir)
        self.max_file_size_bytes = max_file_size_mb * 1024 * 1024
        self.max_file_content_length = max_file_content_length
        self.max_pdf_pages = max_pdf_pages

        # 确保基础目录存在
        self.temp_base_dir.mkdir(parents=True, exist_ok=True)

        # 导入文件工具
        try:
            from utils.file_utils import (
                calculate_file_hash,
                get_file_type,
                encode_to_base64,
                truncate_text,
                sanitize_filename
            )
            self.calculate_file_hash = calculate_file_hash
            self.get_file_type = get_file_type
            self.encode_to_base64 = encode_to_base64
            self.truncate_text = truncate_text
            self.sanitize_filename = sanitize_filename
        except ImportError:
            logger.warning("无法导入文件工具，部分功能可能受限")

    def get_conversation_temp_dir(self, conversation_id: str) -> Path:
        """
        获取对话的临时文件目录

        Args:
            conversation_id: 对话 ID

        Returns:
            Path: 对话临时文件目录
        """
        conv_dir = self.temp_base_dir / conversation_id
        conv_dir.mkdir(parents=True, exist_ok=True)
        return conv_dir

    def save_uploaded_file(
        self,
        conversation_id: str,
        file_content: bytes,
        filename: str,
        ttl_hours: int = 24
    ) -> UploadedFileInfo:
        """
        保存上传的文件到临时目录

        Args:
            conversation_id: 对话 ID
            file_content: 文件内容（二进制）
            filename: 原始文件名
            ttl_hours: 文件保留时间（小时）

        Returns:
            UploadedFileInfo: 上传文件的信息

        Raises:
            ConversationFileManagerError: 保存失败
        """
        try:
            # 验证文件大小
            if len(file_content) > self.max_file_size_bytes:
                size_mb = len(file_content) / (1024 * 1024)
                raise ConversationFileManagerError(
                    f"文件太大: {size_mb:.1f}MB > {self.max_file_size_bytes / (1024 * 1024):.1f}MB"
                )

            # 生成文件哈希（用于去重）
            try:
                file_hash = self.calculate_file_hash(filename) if hasattr(self, 'calculate_file_hash') else None
                if not file_hash:
                    # 如果无法计算文件哈希（文件还不存在），使用内容哈希
                    import hashlib
                    file_hash = hashlib.sha256(file_content).hexdigest()
            except Exception as e:
                logger.warning(f"计算文件哈希失败: {e}，使用内容哈希")
                import hashlib
                file_hash = hashlib.sha256(file_content).hexdigest()

            # 获取文件类型
            try:
                file_type = self.get_file_type(filename) if hasattr(self, 'get_file_type') else self._detect_file_type(filename)
            except:
                file_type = "other"

            # 获取对话的临时目录
            conv_dir = self.get_conversation_temp_dir(conversation_id)

            # 清洁文件名
            try:
                safe_filename = self.sanitize_filename(filename) if hasattr(self, 'sanitize_filename') else filename
            except:
                safe_filename = filename

            # 构建完整的文件路径
            file_path = conv_dir / f"{file_hash}_{safe_filename}"

            # 保存文件
            file_path.write_bytes(file_content)
            logger.info(f"保存上传文件: {file_path}")

            # 生成内容预览
            content_preview = self._generate_preview(file_type, file_content, filename)

            # 计算过期时间
            expires_at = (datetime.now() + timedelta(hours=ttl_hours)).isoformat()

            # 创建文件信息对象
            file_info = UploadedFileInfo(
                file_id=file_hash,
                filename=filename,
                file_type=file_type,
                file_size=len(file_content),
                upload_time=datetime.now().isoformat(),
                content_preview=content_preview,
                metadata={
                    "file_path": str(file_path),
                    "expires_at": expires_at,
                    "conversation_id": conversation_id
                }
            )

            return file_info

        except ConversationFileManagerError:
            raise
        except Exception as e:
            logger.error(f"保存上传文件失败: {e}")
            raise ConversationFileManagerError(f"保存上传文件失败: {e}")

    def extract_file_content(
        self,
        file_path: str
    ) -> str:
        """
        提取文件内容（支持多种格式）

        Args:
            file_path: 文件路径

        Returns:
            str: 提取的文件内容

        Raises:
            ConversationFileManagerError: 提取失败
        """
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"文件不存在: {file_path}")

            file_type = self._detect_file_type(str(file_path))

            # 根据文件类型提取内容
            if file_type == "text":
                return self._extract_text_content(file_path)
            elif file_type == "pdf":
                return self._extract_pdf_content(file_path)
            elif file_type == "word":
                return self._extract_word_content(file_path)
            elif file_type == "excel":
                return self._extract_excel_content(file_path)
            elif file_type == "image":
                return self._extract_image_metadata(file_path)
            elif file_type == "csv":
                return self._extract_csv_content(file_path)
            else:
                return ""

        except Exception as e:
            logger.error(f"提取文件内容失败 ({file_path}): {e}")
            raise ConversationFileManagerError(f"提取文件内容失败: {e}")

    def _extract_text_content(self, file_path: Path) -> str:
        """提取文本文件内容"""
        try:
            content = file_path.read_text(encoding="utf-8", errors="ignore")
            # 截断内容
            return self.truncate_text(content, self.max_file_content_length) if hasattr(self, 'truncate_text') else content[:self.max_file_content_length]
        except Exception as e:
            logger.error(f"提取文本内容失败: {e}")
            return ""

    def _extract_pdf_content(self, file_path: Path) -> str:
        """提取 PDF 文件内容"""
        try:
            # 尝试使用 PyPDF2 库
            try:
                import PyPDF2
                content = []
                with open(file_path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    # 最多提取指定页数
                    max_pages = min(len(reader.pages), self.max_pdf_pages)
                    for i in range(max_pages):
                        page = reader.pages[i]
                        text = page.extract_text()
                        content.append(f"--- 第 {i+1} 页 ---\n{text}")

                full_content = "\n".join(content)
                return self.truncate_text(full_content, self.max_file_content_length) if hasattr(self, 'truncate_text') else full_content[:self.max_file_content_length]
            except ImportError:
                logger.warning("PyPDF2 未安装，无法提取 PDF 内容")
                return "[PDF 文件 - 需要 PyPDF2 库来提取内容]"
        except Exception as e:
            logger.error(f"提取 PDF 内容失败: {e}")
            return f"[PDF 文件提取失败: {str(e)}]"

    def _extract_word_content(self, file_path: Path) -> str:
        """提取 Word 文件内容"""
        try:
            try:
                from docx import Document
                doc = Document(file_path)
                content = "\n".join([para.text for para in doc.paragraphs])
                return self.truncate_text(content, self.max_file_content_length) if hasattr(self, 'truncate_text') else content[:self.max_file_content_length]
            except ImportError:
                logger.warning("python-docx 未安装，无法提取 Word 内容")
                return "[Word 文件 - 需要 python-docx 库来提取内容]"
        except Exception as e:
            logger.error(f"提取 Word 内容失败: {e}")
            return f"[Word 文件提取失败: {str(e)}]"

    def _extract_excel_content(self, file_path: Path) -> str:
        """提取 Excel 文件内容"""
        try:
            try:
                import openpyxl
                import pandas as pd

                # 尝试使用 pandas 读取 Excel 文件（支持 .xlsx 和 .xls）
                try:
                    # 读取所有工作表
                    excel_file = pd.ExcelFile(file_path)
                    all_sheets_content = []

                    for sheet_name in excel_file.sheet_names:
                        # 读取每个工作表
                        df = pd.read_excel(file_path, sheet_name=sheet_name)

                        # 格式化工作表内容
                        sheet_content = f"=== 工作表: {sheet_name} ===\n"

                        # 转换为字符串表格格式
                        sheet_content += df.to_string(index=False, max_rows=100)

                        all_sheets_content.append(sheet_content)

                    # 合并所有工作表内容
                    full_content = "\n\n".join(all_sheets_content)

                    # 截断到最大长度
                    return self.truncate_text(full_content, self.max_file_content_length) if hasattr(self, 'truncate_text') else full_content[:self.max_file_content_length]

                except Exception as pd_error:
                    logger.warning(f"Pandas 读取失败，尝试使用 openpyxl: {pd_error}")

                    # 备用方案：直接使用 openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    all_sheets_content = []

                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        sheet_content = f"=== 工作表: {sheet_name} ===\n"

                        # 读取前100行数据
                        rows_data = []
                        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                            if idx > 100:  # 限制行数
                                break
                            row_str = "\t".join([str(cell) if cell is not None else "" for cell in row])
                            rows_data.append(row_str)

                        sheet_content += "\n".join(rows_data)
                        all_sheets_content.append(sheet_content)

                    wb.close()

                    full_content = "\n\n".join(all_sheets_content)
                    return self.truncate_text(full_content, self.max_file_content_length) if hasattr(self, 'truncate_text') else full_content[:self.max_file_content_length]

            except ImportError as import_err:
                logger.warning(f"Excel 处理库未安装: {import_err}")
                return "[Excel 文件 - 需要 openpyxl 或 pandas 库来提取内容。请运行: pip install openpyxl pandas]"

        except Exception as e:
            logger.error(f"提取 Excel 内容失败: {e}")
            return f"[Excel 文件提取失败: {str(e)}]"

    def _extract_csv_content(self, file_path: Path) -> str:
        """提取 CSV 文件内容"""
        try:
            content = file_path.read_text(encoding="utf-8", errors="ignore")
            # CSV 通常信息较多，截断到较小的长度
            return self.truncate_text(content, self.max_file_content_length // 2) if hasattr(self, 'truncate_text') else content[:self.max_file_content_length // 2]
        except Exception as e:
            logger.error(f"提取 CSV 内容失败: {e}")
            return ""

    def _extract_image_metadata(self, file_path: Path) -> str:
        """
        提取图像信息（返回base64编码 + 元数据，用于多模态LLM）

        返回格式：JSON字符串
        {
            "type": "image",
            "format": "PNG",
            "size": [1920, 1080],
            "base64": "iVBORw0KGgoAAAANSUhEUgAA..."
        }
        """
        try:
            import base64
            import json
            from PIL import Image

            # 读取图片文件并编码为base64
            with open(file_path, "rb") as f:
                image_bytes = f.read()
                base64_str = base64.b64encode(image_bytes).decode('utf-8')

            # 获取图片元数据
            try:
                img = Image.open(file_path)
                img_format = img.format
                img_size = img.size
            except Exception as e:
                logger.warning(f"无法读取图片元数据: {e}")
                img_format = "UNKNOWN"
                img_size = [0, 0]

            # 构建结构化数据
            image_data = {
                "type": "image",
                "format": img_format,
                "size": list(img_size),
                "base64": base64_str,
                "file_path": str(file_path)
            }

            # 返回JSON字符串（带特殊标记，便于后续识别）
            return f"__IMAGE_DATA__{json.dumps(image_data, ensure_ascii=False)}__IMAGE_DATA_END__"

        except Exception as e:
            logger.error(f"提取图像信息失败: {e}")
            return "[图像文件处理失败]"

    def _detect_file_type(self, filename: str) -> str:
        """检测文件类型"""
        suffix = Path(filename).suffix.lower()

        if suffix in ['.txt', '.md', '.markdown']:
            return "text"
        elif suffix == '.pdf':
            return "pdf"
        elif suffix in ['.doc', '.docx']:
            return "word"
        elif suffix in ['.xls', '.xlsx']:
            return "excel"
        elif suffix in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
            return "image"
        elif suffix == '.csv':
            return "csv"
        else:
            return "other"

    def _generate_preview(
        self,
        file_type: str,
        file_content: bytes,
        filename: str
    ) -> str:
        """
        生成文件预览

        Args:
            file_type: 文件类型
            file_content: 文件内容
            filename: 文件名

        Returns:
            str: 预览内容（文本或 base64）
        """
        try:
            if file_type == "text":
                # 文本文件：显示前 200 个字符
                text = file_content.decode("utf-8", errors="ignore")
                return text[:200] + "..." if len(text) > 200 else text
            elif file_type == "image":
                # 图像文件：编码为 base64
                try:
                    return self.encode_to_base64(filename) if hasattr(self, 'encode_to_base64') else base64.b64encode(file_content).decode("utf-8")
                except:
                    return f"[图像文件: {filename}]"
            else:
                # 其他文件：显示文件信息
                size_kb = len(file_content) / 1024
                return f"[{file_type.upper()} 文件: {filename}, {size_kb:.1f}KB]"
        except Exception as e:
            logger.warning(f"生成预览失败: {e}")
            return f"[{file_type} 文件]"

    def cleanup_conversation_files(self, conversation_id: str) -> int:
        """
        清理对话的所有临时文件

        Args:
            conversation_id: 对话 ID

        Returns:
            int: 删除的文件数
        """
        try:
            conv_dir = self.temp_base_dir / conversation_id
            if not conv_dir.exists():
                return 0

            deleted_count = 0
            for file_path in conv_dir.rglob("*"):
                if file_path.is_file():
                    try:
                        file_path.unlink()
                        deleted_count += 1
                        logger.info(f"删除对话文件: {file_path}")
                    except Exception as e:
                        logger.error(f"删除文件失败 ({file_path}): {e}")

            # 删除对话目录
            try:
                import shutil
                if conv_dir.exists():
                    shutil.rmtree(conv_dir)
            except Exception as e:
                logger.error(f"删除对话目录失败: {e}")

            logger.info(f"清理对话文件完成: 删除 {deleted_count} 个文件")
            return deleted_count

        except Exception as e:
            logger.error(f"清理对话文件失败: {e}")
            raise ConversationFileManagerError(f"清理对话文件失败: {e}")
